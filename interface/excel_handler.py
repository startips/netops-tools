#!/usr/bin/python3
# -*- coding: utf-8 -*-

"""
excel_handler.py - Excel 表格处理模块

仅支持 .xlsx 格式，提供 Excel 读写、样式设置功能。
"""

import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font


class excel:  # Excel表格处理 只支持.xlsx格式
    def __init__(self, filename):  # 初始化
        self.filename = filename  # 文件名 .xlsx

    def _mark_highlight(self, ws, keywords=None, start_row=2, color="FFFF0000"):
        """
        标记含有关键字的单元格
        
        Args:
            ws: 工作表对象
            keywords: 关键字列表，默认 ['未通过']
            start_row: 检查起始行，默认2（跳过标题行）
            color: 标记颜色，默认红色
        """
        if keywords is None:
            keywords = ['未通过']
        highlight_font = Font(color=color)
        for row in ws.iter_rows(min_row=start_row):
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                for key in keywords:
                    if isinstance(value, str) and key in value:
                        cell.font = highlight_font
                        break

    # 写入数据
    def excel_write(self, title, data, sheetname='data01', sheetIndex=1, highlight=False):  # 一次性写入 data 格式为[[],[]]
        self.wb_obj = Workbook()
        self.wb_obj.active
        title_local = title  # 标题 list
        data_local = data
        sheetname_local = sheetname  # sheet名称
        sheetIndex_local = sheetIndex - 1  # sheet的位置 默认是第一张表 位置从0开始
        wsObj = self.wb_obj.create_sheet(sheetname_local, sheetIndex_local)
        wsObj.append(title_local)
        title_font = Font(b='bold', size='12')
        for i in range(1, len(title) + 1):
            cell = wsObj.cell(row=1, column=i)
            cell.font = title_font
        for row_data in data_local:
            try:
                wsObj.append(row_data)  # 写入数据
            except Exception:
                for row in row_data:
                    wsObj.append(row)
        if highlight:
            self._mark_highlight(wsObj)

    def excel_write_multi_sheet(self, sheets_data, highlight=False):
        """
        写入多个sheet到一个Excel文件（只组装数据，由 save_file() 统一保存）
        
        Args:
            sheets_data: list of dict, 每个dict包含:
                - title: list, 标题行
                - data: list of list, 数据行
                - sheetname: str, sheet名称
                - highlight: bool, 是否高亮标记（可选，默认跟随函数参数）
            highlight: bool, 全局默认是否高亮标记，默认False
        
        示例:
            sheets_data = [
                {'title': ['列1', '列2'], 'data': [['a', 'b'], ['c', 'd']], 'sheetname': 'Sheet1'},
                {'title': ['列A', '列B'], 'data': [['x', 'y']], 'sheetname': 'Sheet2', 'highlight': True},
            ]
        """
        self.wb_obj = Workbook()
        # 删除默认创建的 sheet
        if 'Sheet' in self.wb_obj.sheetnames:
            del self.wb_obj['Sheet']
        
        title_font = Font(b='bold', size='12')
        
        for idx, sheet_info in enumerate(sheets_data):
            title = sheet_info.get('title', [])
            data = sheet_info.get('data', [])
            sheetname = sheet_info.get('sheetname', f'Sheet{idx + 1}')
            # 优先用 sheet 自己的配置，没有就用函数参数的默认值
            sheet_highlight = sheet_info.get('highlight', highlight)
            
            # 创建新 sheet
            ws = self.wb_obj.create_sheet(sheetname, idx)
            
            # 写入标题
            if title:
                ws.append(title)
                for i in range(1, len(title) + 1):
                    cell = ws.cell(row=1, column=i)
                    cell.font = title_font
            
            # 写入数据
            for row_data in data:
                try:
                    ws.append(row_data)
                except Exception:
                    for row in row_data:
                        ws.append(row)
            
            # 标记颜色（按 sheet 单独控制）
            if sheet_highlight:
                self._mark_highlight(ws)

    def excel_creat(self, title, sheetname='data01', sheetIndex=1):  # 创建对象并设置好列头
        self.wb_obj = Workbook()
        self.wb_obj.active
        title_local = title  # 标题 list
        sheetname_local = sheetname  # sheet名称
        sheetIndex_local = sheetIndex - 1  # sheet的位置 默认是第一张表 位置从0开始
        self.wsobj = self.wb_obj.create_sheet(sheetname_local, sheetIndex_local)
        self.wsobj.append(title_local)
        title_font = Font(b='bold', size='12')
        for i in range(1, len(title) + 1):
            cell = self.wsobj.cell(row=1, column=i)
            cell.font = title_font

    def write_row(self, data: list):  # 写入单行数据[]
        row_data = data
        self.wsobj.append(row_data)

    # 保存文件
    def save_file(self):  # 保存文件
        timeNow = time.strftime('%Y-%m-%d_%H%M%S', time.localtime(time.time()))
        filename_save = self.filename.replace('.xlsx', '')
        filename_save = f'{filename_save}_{timeNow}.xlsx'  # 返回文件名
        self.wb_obj.save(filename_save)  # 存盘
        self.wb_obj.close()  # 关闭
        return filename_save

    # 读取数据 默认打开第一个sheet从第二行读
    def excel_read(self, sheet=1, row=0, column=0, row_start=2, column_start=1):
        file_local = self.filename
        wb = load_workbook(filename=file_local, data_only=True)  # 打开一个excel对象
        sheetnames = wb.sheetnames  # 获取sheets
        if isinstance(sheet, int):
            ws = wb[sheetnames[sheet - 1]]  # 按序号（从1开始）
        else:
            ws = wb[sheet]  # 按名字
        row_start_local = row_start  # 起始行
        column_start_local = column_start  # 起始列
        if row == 0:
            row = ws.max_row  # 最大行
        if column == 0:
            column = ws.max_column  # 最大列
        data_result = []  # 结果集
        for rx in range(row_start_local, row + 1):  # 循环读取每一行sheet数据
            info = []  # 一行的数据集
            for cx in range(column_start_local, column + 1):  # 循环读取每一行的列数据
                cell_info = ws.cell(row=rx, column=cx).value
                info.append(cell_info)
            data_result.append(info)
        wb.close()  # 关闭
        return data_result

    def excelReadCread(self, ):  # 打开excel对象
        file_local = self.filename
        self.wb = load_workbook(filename=file_local)  # 打开一个excel对象
        sheetNums = self.wb.sheetnames
        return sheetNums

    def excelReadSheet(self, sheet=1, row=0, column=0, row_start=2, column_start=1):  # 读取一个sheet
        sheetnames = self.wb.sheetnames  # 获取sheets
        if isinstance(sheet, int):
            ws = self.wb[sheetnames[sheet - 1]]  # 按序号（从1开始）
        else:
            ws = self.wb[sheet]  # 按名字
        row_start_local = row_start  # 起始行
        column_start_local = column_start  # 起始列
        if row == 0:
            row = ws.max_row  # 行计数
        if column == 0:
            column = ws.max_column  # 列计数
        data_result = []  # 结果集
        for rx in range(row_start_local, row + 1):  # 循环读取每一行sheet数据
            info = []  # 一行的数据集
            for cx in range(column_start_local, column + 1):  # 循环读取每一行的列数据
                cell_info = ws.cell(row=rx, column=cx).value
                info.append(cell_info)
            data_result.append(info)
        return data_result

    def excelClose(self):
        self.wb.close()  # 关闭
