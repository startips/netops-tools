#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""
device_diff_compare.py - 设备巡检数据差异对比

对比两个时期的巡检成功数据，找出差异设备，并从原始数据汇总中查找原因。

用法：
    python device_diff_compare.py <数据A> <数据B> <原始数据> [输出文件]
    python device_diff_compare.py  # 交互模式

示例：
    python device_diff_compare.py \\
        "2026Q2/巡检成功数据汇总_2026-06-15_141825.xlsx" \\
        "2026五一节前巡检/巡检成功数据汇总_2026-04-24_111034.xlsx" \\
        "2026Q2/巡检原始数据汇总_2026-06-09_175856.xlsx" \\
        "设备差异对比_Q2vs五一.xlsx"
"""

from typing import Optional
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
import os
import sys
import logging

logger = logging.getLogger(__name__)


def load_device_data(filepath: str) -> dict:
    """读取巡检成功数据，返回 {IP: {name, group, type}} 字典"""
    wb = openpyxl.load_workbook(filepath)
    ws: Worksheet = wb["巡检网元汇总"]  # type: ignore
    
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        name, group, type_, version, patch, scene, ip, esn = row
        if ip:
            ip = str(ip).strip()
            data[ip] = {"name": name, "group": group, "type": type_, "ip": ip}
    
    wb.close()
    logger.info("读取 '%s' 成功, %d 个IP", os.path.basename(filepath), len(data))
    return data


def load_raw_data(filepath: str) -> dict:
    """读取原始数据汇总，返回IP信息字典"""
    wb = openpyxl.load_workbook(filepath)
    ws: Worksheet = wb["巡检原始数据"]  # type: ignore
    
    all_ips: set = set()
    has_note: dict = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        ip, user, ab, note, time = row
        if ip:
            ip = str(ip).strip()
            all_ips.add(ip)
            if note or time:
                has_note[ip] = (note, time)
    
    wb.close()
    logger.info("读取原始数据 '%s' 成功, %d 个IP, %d 个有备注",
                os.path.basename(filepath), len(all_ips), len(has_note))
    return {"all_ips": all_ips, "has_note": has_note}


def get_reason(ip: str, raw_data: dict) -> str:
    """根据IP从原始数据中获取原因"""
    if ip in raw_data["has_note"]:
        note, time = raw_data["has_note"][ip]
        parts = []
        if note:
            parts.append(str(note))
        if time:
            parts.append(str(time))
        return " | ".join(parts) if parts else "无记录"
    elif ip in raw_data["all_ips"]:
        return "匹配到但无备注"
    else:
        return "匹配不到"


def compare_devices(data_a: dict, data_b: dict, raw_data: dict) -> tuple:
    """对比两组设备数据，返回差异列表"""
    ips_a = set(data_a.keys())
    ips_b = set(data_b.keys())
    
    # A有B没有
    missing_in_b = {ip: data_a[ip] for ip in ips_a if ip not in ips_b}
    # B有A没有
    missing_in_a = {ip: data_b[ip] for ip in ips_b if ip not in ips_a}
    
    logger.info("差异统计: A有B缺失 %d, B有A新增 %d", len(missing_in_b), len(missing_in_a))
    return missing_in_b, missing_in_a


def export_excel(missing_in_b: dict, missing_in_a: dict, raw_data: dict, 
                 output_path: str, label_a: str, label_b: str) -> int:
    """导出差异对比结果到Excel"""
    out_wb = openpyxl.Workbook()
    out_ws: Worksheet = out_wb.active  # type: ignore
    out_ws.title = "设备差异对比"
    
    # 样式
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fill_no_record = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    fill_no_note = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    
    # 表头
    headers = ["序号", "对比类型", "网元名称", "网元分组", "网元类型", "网元IP", "原因"]
    for col, header in enumerate(headers, 1):
        cell = out_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center")
    
    # 写入数据
    idx = 0
    
    # A有B没有
    label_missing = f"{label_a}有{label_b}缺失"
    for ip in sorted(missing_in_b.keys()):
        idx += 1
        data = missing_in_b[ip]
        reason = get_reason(ip, raw_data)
        row_data = [idx, label_missing, data["name"], data["group"], data["type"], ip, reason]
        
        row_fill = None
        if reason == "匹配不到":
            row_fill = fill_no_record
        elif reason == "匹配到但无备注":
            row_fill = fill_no_note
        
        for col, value in enumerate(row_data, 1):
            cell = out_ws.cell(row=idx+1, column=col, value=value)
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            if row_fill:
                cell.fill = row_fill
    
    # B有A没有
    label_new = f"{label_a}无{label_b}新增"
    for ip in sorted(missing_in_a.keys()):
        idx += 1
        data = missing_in_a[ip]
        reason = get_reason(ip, raw_data)
        row_data = [idx, label_new, data["name"], data["group"], data["type"], ip, reason]
        
        row_fill = None
        if reason == "匹配不到":
            row_fill = fill_no_record
        elif reason == "匹配到但无备注":
            row_fill = fill_no_note
        
        for col, value in enumerate(row_data, 1):
            cell = out_ws.cell(row=idx+1, column=col, value=value)
            cell.border = thin_border
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            if row_fill:
                cell.fill = row_fill
    
    # 调整列宽
    out_ws.column_dimensions["A"].width = 8
    out_ws.column_dimensions["B"].width = 16
    out_ws.column_dimensions["C"].width = 38
    out_ws.column_dimensions["D"].width = 15
    out_ws.column_dimensions["E"].width = 22
    out_ws.column_dimensions["F"].width = 18
    out_ws.column_dimensions["G"].width = 45
    
    # 冻结首行
    out_ws.freeze_panes = "A2"
    
    # 添加自动筛选
    out_ws.auto_filter.ref = f"A1:G{idx+1}"
    
    # 保存
    out_wb.save(output_path)
    logger.info("导出完成: %s, %d 条记录", output_path, idx)
    return idx


def extract_label(filepath: str) -> str:
    """从文件路径提取友好标签，如 '2026Q2' → 'Q2', '2026五一节前巡检' → '五一'"""
    # 获取上层目录名或文件名
    basename = os.path.basename(filepath)
    dirname = os.path.basename(os.path.dirname(filepath))
    
    # 常见映射
    label_map = {
        "2026Q2": "Q2",
        "2026Q1": "Q1",
        "2026五一节前巡检": "五一",
        "2025年底巡检": "年底",
    }
    
    # 先从目录名匹配
    for key, label in label_map.items():
        if key in dirname:
            return label
    
    # 从文件名匹配
    for key, label in label_map.items():
        if key in basename:
            return label
    
    # 都没匹配到，用目录名
    return dirname if dirname else "数据"


def main(data_a_path: str, data_b_path: str, raw_data_path: str, output_path: Optional[str] = None) -> str:
    """主函数"""
    # 提取标签
    label_a = extract_label(data_a_path)
    label_b = extract_label(data_b_path)
    
    # 默认输出文件名
    if not output_path:
        output_dir = os.path.dirname(data_a_path)
        output_path = os.path.join(output_dir, f"设备差异对比_{label_a}vs{label_b}.xlsx")
    
    # 读取数据
    data_a = load_device_data(data_a_path)
    data_b = load_device_data(data_b_path)
    raw_data = load_raw_data(raw_data_path)
    
    # 对比
    missing_in_b, missing_in_a = compare_devices(data_a, data_b, raw_data)
    
    # 导出
    total = export_excel(missing_in_b, missing_in_a, raw_data, output_path, label_a, label_b)
    
    # 统计结果
    print(f"\n✅ 完成: {len(data_a)} vs {len(data_b)} 个IP → {total} 条差异记录")
    print(f"   {label_a}有{label_b}缺失: {len(missing_in_b)} 台")
    print(f"   {label_a}无{label_b}新增: {len(missing_in_a)} 台")
    print(f"📁 输出文件: {output_path}")
    
    return output_path


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format='%(message)s')
    
    if len(sys.argv) >= 4:
        # 命令行模式
        data_a = sys.argv[1]
        data_b = sys.argv[2]
        raw_data = sys.argv[3]
        output = sys.argv[4] if len(sys.argv) > 4 else None
        main(data_a, data_b, raw_data, output)
    else:
        # 交互模式
        print("📊 设备巡检数据差异对比工具")
        print("=" * 50)
        
        data_a_path = input("请输入数据A路径（巡检成功数据）: ").strip().strip('"')
        data_b_path = input("请输入数据B路径（巡检成功数据）: ").strip().strip('"')
        raw_data_path = input("请输入原始数据汇总路径: ").strip().strip('"')
        output_path = input("请输入输出文件路径（回车自动生成）: ").strip().strip('"') or None
        
        if not os.path.exists(data_a_path):
            print(f"❌ 文件不存在: {data_a_path}")
            sys.exit(1)
        if not os.path.exists(data_b_path):
            print(f"❌ 文件不存在: {data_b_path}")
            sys.exit(1)
        if not os.path.exists(raw_data_path):
            print(f"❌ 文件不存在: {raw_data_path}")
            sys.exit(1)
        
        main(data_a_path, data_b_path, raw_data_path, output_path)
